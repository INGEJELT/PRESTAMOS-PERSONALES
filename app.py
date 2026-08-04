import sqlite3
import os
import pandas as pd
from flask import Flask, render_template, request, redirect, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import re
import uuid
import json
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.secret_key = "cardenal_master_key_2026"
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
DB_NAME = "cardenal_napoles.db"

# --- NUEVO: Diccionario para controlar sesiones activas ---
sesiones_activas = {}
app.secret_key = "cardenal_master_key_2026"
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
DB_NAME = "cardenal_napoles.db"

# ------------------------------------------------------------
# 1. INICIALIZACIÓN DE BASE DE DATOS
# ------------------------------------------------------------
def init_db():
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password TEXT, rol TEXT, sucursal TEXT)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS prestamos (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nomina TEXT, empleado TEXT NOT NULL,
        area TEXT, monto_inicial REAL, saldo_pendiente REAL, fecha_otorgamiento TEXT, sucursal TEXT)''')
    
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN semana_otorgada INTEGER')
    except: pass
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN autoriza TEXT')
    except: pass
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN motivo_adicional TEXT')
    except: pass
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN id_empleado TEXT')
    except: pass
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN fecha_ingreso TEXT')
    except: pass
    try: cursor.execute('ALTER TABLE prestamos ADD COLUMN puesto TEXT')
    except: pass

    cursor.execute('''CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT, prestamo_id INTEGER, fecha TEXT, semana INTEGER, tipo TEXT, monto REAL,
        FOREIGN KEY(prestamo_id) REFERENCES prestamos(id))''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS arqueos (
        id INTEGER PRIMARY KEY AUTOINCREMENT, sucursal TEXT, fecha TEXT, semana INTEGER, 
        fondo_sistema REAL, efectivo_real REAL, diferencia REAL, observaciones TEXT, usuario TEXT, detalle TEXT)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS auditoria (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha_hora TEXT, usuario TEXT, accion TEXT,
        empleado_nombre TEXT, detalle TEXT, motivo TEXT, sucursal TEXT)''')
    
    # Agregar columna detalle a arqueos si no existe
    try:
        cursor.execute('ALTER TABLE arqueos ADD COLUMN detalle TEXT')
    except:
        pass
    
    sucursales = ['TACUBA', 'BOMBILLA', 'RESTORANES', 'NAPOLES', 'BRIGAR', 'BGARI']
    for s in sucursales:
        cursor.execute('INSERT OR IGNORE INTO cajas (sucursal, saldo_inicial, saldo_actual) VALUES (?, 0, 0)', (s,))
    
    admin_exist = cursor.execute('SELECT * FROM usuarios WHERE username="SISTEMAS"').fetchone()
    if not admin_exist:
        pw = generate_password_hash('admin123')
        cursor.execute('INSERT INTO usuarios (username, password, rol, sucursal) VALUES (?,?,?,?)', ('SISTEMAS', pw, 'ADMIN', 'TODAS'))
        
    conn.commit()
    conn.close()

def es_admin():
    return session.get('rol') == 'ADMIN'

# ------------------------------------------------------------
# CONTROL DE SESIONES ACTIVAS Y EXPULSIÓN
# ------------------------------------------------------------
@app.before_request
def verificar_sesion_activa():
    if request.path.startswith('/static/') or request.path == '/login':
        return None
        
    if 'user' in session:
        user = session['user']
        token_actual = session.get('session_token')
        
        # Si el usuario no está en el diccionario de sesiones activas (porque fue expulsado)
        if user not in sesiones_activas:
            session.clear()
            flash("Tu sesión ha sido cerrada remotamente por el administrador.", "danger")
            return redirect('/login')
        
        # Si está, validamos token
        token_valido = sesiones_activas[user].get('token')
        if token_valido != token_actual and user != 'SISTEMAS':
            session.clear()
            flash("Tu sesión ha sido cerrada remotamente por el administrador.", "danger")
            return redirect('/login')
        else:
            # Actualizar última actividad
            sesiones_activas[user]['ultima_actividad'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Si no hay usuario en sesión, no hacemos nada

# ------------------------------------------------------------
# BLOQUEO OBLIGATORIO DE ARQUEO SEMANAL
# ------------------------------------------------------------
@app.before_request
def forzar_arqueo_semanal():
    rutas_permitidas = ['/login', '/logout', '/guardar_arqueo']
    if request.path in rutas_permitidas or request.path.startswith('/static/'):
        return None
    
    if 'user' not in session:
        return None
    
    sucursal = session.get('sucursal')
    if not sucursal or sucursal == 'TODAS': 
        return None
    
    hoy = datetime.now()
    semana_actual = hoy.isocalendar()[1]
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    arqueo_hecho = conn.execute('SELECT id FROM arqueos WHERE sucursal = ? AND semana = ? AND strftime("%Y", fecha) = ?', 
                                (sucursal, semana_actual, str(hoy.year))).fetchone()
    conn.close()
    
    if not arqueo_hecho and not request.path.startswith(f'/resumen_empresa/{sucursal}'):
        flash("¡ARQUEO SEMANAL PENDIENTE! Es obligatorio realizar el cuadre de caja de esta semana antes de continuar operando.", "warning")
        return redirect(f'/resumen_empresa/{sucursal}')

# ------------------------------------------------------------
# PREFIJOS PERSONALIZADOS PARA ID
# ------------------------------------------------------------
def obtener_prefijo_sucursal(sucursal):
    prefijos = {
        'TACUBA': 'TAC',
        'BRIGAR': 'BRI',
        'BOMBILLA': 'CDB',
        'RESTORANES': 'RIN',
        'NAPOLES': 'CNA',
        'BGARI': 'BGG'
    }
    return prefijos.get(sucursal.upper(), 'XXX')

def formatear_nomina_para_id(nomina):
    """Convierte la nómina a un número con ceros a la izquierda según su longitud."""
    try:
        # Convertir a entero para eliminar ceros a la izquierda
        num = int(str(nomina).strip())
    except ValueError:
        # Si no es numérico, devolver como está (ej: "A123")
        return str(nomina).strip()
    
    # Determinar cantidad de ceros según los dígitos del número
    digitos = len(str(num))
    if digitos <= 2:
        ceros = 3  # 000
    elif digitos == 3:
        ceros = 2  # 00
    elif digitos == 4:
        ceros = 1  # 0
    else:
        ceros = 0  # sin ceros
    
    return str(num).zfill(ceros + digitos)  # zfill rellena con ceros a la izquierda

def generar_id_empleado(sucursal, nomina, conn):
    """Genera un ID de empleado con el formato: PREFIJO + CEROS + NOMINA [+ _N]"""
    prefijo = obtener_prefijo_sucursal(sucursal)
    nomina_formateada = formatear_nomina_para_id(nomina)
    base_id = f"{prefijo}{nomina_formateada}"
    
    cursor = conn.cursor()
    # Verificar si el ID base ya existe
    cursor.execute("SELECT id_empleado FROM prestamos WHERE sucursal = ? AND id_empleado = ?", (sucursal, base_id))
    if not cursor.fetchone():
        return base_id
    
    # Si existe, agregar sufijo _2, _3, ...
    contador = 2
    while True:
        nuevo_id = f"{base_id}_{contador}"
        cursor.execute("SELECT id_empleado FROM prestamos WHERE sucursal = ? AND id_empleado = ?", (sucursal, nuevo_id))
        if not cursor.fetchone():
            return nuevo_id
        contador += 1

# ------------------------------------------------------------
# 2. CONFIGURACIÓN DE FONDO
# ------------------------------------------------------------
@app.route('/configurar_fondo', methods=['POST'])
def configurar_fondo():
    if not es_admin():
        return redirect('/')
    sucursal = request.form.get('sucursal_caja')
    monto = int(float(request.form.get('monto_inicial')))
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.execute('UPDATE cajas SET saldo_inicial = ?, saldo_actual = ? WHERE sucursal = ?', (monto, monto, sucursal))
    conn.commit()
    conn.close()
    flash(f"Fondo de {sucursal} actualizado.", "success")
    return redirect('/')

# ------------------------------------------------------------
# Sincronizar datos de empleados desde archivos Excel
# ------------------------------------------------------------
@app.route('/sincronizar_empleados', methods=['POST'])
def sincronizar_empleados():
    if not es_admin():
        return redirect('/')
    
    sucursal = request.form.get('sucursal_sync', 'TODAS')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    try:
        cursor.execute('ALTER TABLE prestamos ADD COLUMN puesto TEXT')
    except:
        pass
    
    if sucursal == 'TODAS':
        prestamos = cursor.execute('SELECT id, nomina, sucursal FROM prestamos').fetchall()
    else:
        prestamos = cursor.execute('SELECT id, nomina, sucursal FROM prestamos WHERE sucursal = ?', (sucursal,)).fetchall()
    
    actualizados = 0
    errores = 0
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(base_dir, 'empleados')
    
    if not os.path.exists(folder_path):
        flash(f"La carpeta 'empleados' no existe en {folder_path}", "danger")
        return redirect(request.referrer or '/')
    
    for prestamo in prestamos:
        p_id, nomina_db, suc = prestamo
        
        try:
            nomina_normalizada = str(int(float(nomina_db)))
        except:
            nomina_normalizada = str(nomina_db).strip()
        
        archivo_encontrado = None
        for nombre in os.listdir(folder_path):
            if nombre.upper().startswith(f"DEPARTAMENTO {suc.upper()}") or nombre.upper().startswith(f"DEPARTAMENTO_{suc.upper()}"):
                archivo_encontrado = os.path.join(folder_path, nombre)
                break
        if not archivo_encontrado:
            for nombre in os.listdir(folder_path):
                if nombre.upper().startswith(suc.upper()):
                    archivo_encontrado = os.path.join(folder_path, nombre)
                    break
        
        if not archivo_encontrado:
            print(f"No se encontró archivo para sucursal {suc}")
            errores += 1
            continue
        
        try:
            df_headers = pd.read_excel(archivo_encontrado, nrows=1, header=None)
            num_cols = len(df_headers.columns)
            fecha_col_idx = 5 if num_cols >= 6 else 3
            puesto_col_idx = 4 if num_cols >= 6 else 2
            
            df = pd.read_excel(archivo_encontrado, header=None, dtype=str)
            departamento_actual = ""
            encontrado = False
            for idx, row in df.iterrows():
                primera_col = str(row[0]) if pd.notna(row[0]) else ""
                if "Departamento:" in primera_col:
                    if "/" in primera_col:
                        partes = primera_col.split("/")
                        if len(partes) >= 2:
                            departamento_actual = partes[1].strip()
                        else:
                            departamento_actual = primera_col.replace("Departamento:", "").strip()
                    else:
                        departamento_actual = primera_col.replace("Departamento:", "").strip()
                    continue
                
                clave_excel_raw = str(row[0]).strip() if pd.notna(row[0]) else ""
                if not clave_excel_raw or clave_excel_raw == "nan" or clave_excel_raw == "Clave":
                    continue
                
                try:
                    clave_normalizada = str(int(float(clave_excel_raw)))
                except:
                    clave_normalizada = clave_excel_raw
                
                if clave_normalizada == nomina_normalizada:
                    puesto = ""
                    if len(row) > puesto_col_idx and pd.notna(row[puesto_col_idx]):
                        puesto = str(row[puesto_col_idx]).strip()
                    
                    fecha_alta_raw = None
                    if len(row) > fecha_col_idx and pd.notna(row[fecha_col_idx]):
                        fecha_alta_raw = row[fecha_col_idx]
                    
                    fecha_alta = ""
                    if fecha_alta_raw:
                        fecha_str = str(fecha_alta_raw).strip()
                        try:
                            fecha_str = fecha_str.replace('-', '/')
                            fecha_alta = datetime.strptime(fecha_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                        except:
                            fecha_alta = fecha_str
                    
                    cursor.execute('UPDATE prestamos SET area = ?, puesto = ?, fecha_ingreso = ? WHERE id = ?',
                                   (departamento_actual, puesto, fecha_alta, p_id))
                    actualizados += 1
                    encontrado = True
                    break
            if not encontrado:
                print(f"Empleado con nómina {nomina_db} (normalizada: {nomina_normalizada}) no encontrado en {archivo_encontrado}")
                errores += 1
        except Exception as e:
            print(f"Error procesando {archivo_encontrado}: {e}")
            errores += 1
    
    conn.commit()
    conn.close()
    
    flash(f"Sincronización completada. {actualizados} empleados actualizados (puesto, departamento, fecha ingreso), {errores} errores.", "success")
    return redirect(request.referrer or '/')

# ------------------------------------------------------------
# 3. DASHBOARD MAESTRO
# ------------------------------------------------------------
@app.route('/')
def inicio():
    return redirect(f'/semana/{datetime.now().isocalendar()[1]}/sucursal/TODAS')

@app.route('/semana/<int:num_sem>/sucursal/<string:suc>')
def index(num_sem, suc):
    if 'user' not in session:
        return redirect('/login')
    if not es_admin():
        suc = session['sucursal']
        
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    todas_cajas = conn.execute('SELECT sucursal, saldo_inicial FROM cajas').fetchall()
    for caja in todas_cajas:
        deuda_total = conn.execute('SELECT SUM(saldo_pendiente) FROM prestamos WHERE sucursal = ?', (caja['sucursal'],)).fetchone()[0] or 0
        saldo_real = caja['saldo_inicial'] - deuda_total
        conn.execute('UPDATE cajas SET saldo_actual = ? WHERE sucursal = ?', (saldo_real, caja['sucursal']))
    conn.commit()

    cajas = conn.execute('SELECT * FROM cajas').fetchall()
    res_anio = conn.execute('SELECT fecha_otorgamiento FROM prestamos ORDER BY id DESC LIMIT 1').fetchone()
    anio_act = str(datetime.now().year) 
    if res_anio and res_anio[0]:
        match = re.search(r'\d{4}', str(res_anio[0]))
        if match:
            anio_act = match.group()
    
    if suc == 'TODAS':
        stats = conn.execute('SELECT SUM(monto_inicial), SUM(saldo_pendiente) FROM prestamos WHERE saldo_pendiente > 0').fetchone()
        fondo_inicial_total = conn.execute('SELECT SUM(saldo_inicial) FROM cajas').fetchone()[0] or 0
    else:
        stats = conn.execute('SELECT SUM(monto_inicial), SUM(saldo_pendiente) FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (suc,)).fetchone()
        fondo_inicial_total = conn.execute('SELECT SUM(saldo_inicial) FROM cajas WHERE sucursal = ?', (suc,)).fetchone()[0] or 0

    saldo_global_historico = stats[1] or 0
    fondo_actual = fondo_inicial_total - saldo_global_historico

    q_rec = "SELECT SUM(m.monto) FROM movimientos m JOIN prestamos p ON m.prestamo_id = p.id WHERE m.semana = ? AND m.tipo = 'ABONO'"
    q_sal = "SELECT SUM(saldo_pendiente) FROM prestamos WHERE id IN (SELECT prestamo_id FROM movimientos WHERE semana = ?)"
    p_sem = [num_sem]
    
    if suc != 'TODAS':
        q_rec += " AND p.sucursal = ?"
        q_sal += " AND sucursal = ?"
        p_sem.append(suc)
        
    tot_rec = conn.execute(q_rec, p_sem).fetchone()[0] or 0
    tot_sal = conn.execute(q_sal, p_sem).fetchone()[0] or 0
    totales_semana = (tot_sal, tot_rec)

    usuarios_lista = conn.execute('SELECT * FROM usuarios WHERE username != "SISTEMAS"').fetchall() if es_admin() else []

    if suc == 'TODAS':
        query = '''SELECT p.*, 
                          IFNULL(p.semana_otorgada, strftime('%W', p.fecha_otorgamiento)) as semana_otorgamiento, 
                          (SELECT MAX(semana) FROM movimientos WHERE prestamo_id = p.id AND semana = ?) as ultima_semana, 
                          (SELECT SUM(monto) FROM movimientos WHERE prestamo_id = p.id AND semana = ? AND tipo = 'ABONO') as abono_semana 
                   FROM prestamos p'''
        params = [num_sem, num_sem]
    else:
        query = '''SELECT p.*, 
                          IFNULL(p.semana_otorgada, strftime('%W', p.fecha_otorgamiento)) as semana_otorgamiento, 
                          (SELECT MAX(semana) FROM movimientos WHERE prestamo_id = p.id AND semana = ?) as ultima_semana, 
                          (SELECT SUM(monto) FROM movimientos WHERE prestamo_id = p.id AND semana = ? AND tipo = 'ABONO') as abono_semana 
                   FROM prestamos p 
                   WHERE sucursal = ?'''
        params = [num_sem, num_sem, suc]
        
    empleados = conn.execute(query, params).fetchall()
    
    ultimos_arqueos_dict = {}
    if es_admin():
        sucursales_unicas = conn.execute('SELECT DISTINCT sucursal FROM arqueos').fetchall()
        for s in sucursales_unicas:
            ult = conn.execute('SELECT * FROM arqueos WHERE sucursal = ? ORDER BY fecha DESC LIMIT 1', (s['sucursal'],)).fetchone()
            if ult:
                ultimos_arqueos_dict[s['sucursal']] = dict(ult)
    else:
        ult = conn.execute('SELECT * FROM arqueos WHERE sucursal = ? ORDER BY fecha DESC LIMIT 1', (session['sucursal'],)).fetchone()
        if ult:
            ultimos_arqueos_dict[session['sucursal']] = dict(ult)
    
    conn.close()
    
    semana_real = datetime.now().isocalendar()[1]
    
    return render_template('index.html',
                          empleados=empleados,
                          stats=stats,
                          totales_semana=totales_semana,
                          fondo_actual=fondo_actual,
                          fondo_inicial_total=fondo_inicial_total,
                          semana_act=num_sem,
                          suc_act=suc,
                          cajas=cajas,
                          anio_act=anio_act,
                          admin=es_admin(),
                          usuarios=usuarios_lista,
                          sucursal_usuario=session.get('sucursal', 'TODAS'),
                          datetime=datetime,
                          semana_real=semana_real,
                          ultimos_arqueos_dict=ultimos_arqueos_dict)

# ------------------------------------------------------------
# EDICIÓN DIRECTA DE PRÉSTAMOS
# ------------------------------------------------------------
@app.route('/editar_prestamo_maestro', methods=['POST'])
def editar_prestamo_maestro():
    if not es_admin():
        return redirect('/')
    
    p_id = request.form.get('prestamo_id')
    nueva_nomina = request.form.get('nomina')
    nuevo_nombre = request.form.get('empleado')
    nueva_area = request.form.get('area')
    nuevo_monto = int(float(request.form.get('monto_inicial')))
    nuevo_autoriza = request.form.get('autoriza')
    nueva_fecha_ingreso = request.form.get('fecha_ingreso')
    nueva_fecha_otorgamiento = request.form.get('fecha_otorgamiento')

    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    viejo = cursor.execute('SELECT * FROM prestamos WHERE id = ?', (p_id,)).fetchone()
    if not viejo:
        conn.close()
        flash("Préstamo no encontrado.", "danger")
        return redirect(request.referrer or '/')

    monto_viejo = viejo[4]
    sucursal = viejo[7]
    id_actual = viejo[9] if len(viejo) > 9 else None
    diferencia = nuevo_monto - monto_viejo

    prefijo_esperado = obtener_prefijo_sucursal(sucursal)
    if not id_actual or id_actual.strip() == '' or id_actual == 'N/A' or id_actual == 'ENLACE' or not id_actual.startswith(prefijo_esperado):
        nuevo_id = generar_id_empleado(sucursal, nueva_nomina, conn)
    else:
        nuevo_id = id_actual

    cursor.execute('''UPDATE prestamos 
                      SET nomina=?, empleado=?, area=?, monto_inicial=?, 
                          saldo_pendiente = round(saldo_pendiente + ?, 2), 
                          autoriza=?, id_empleado=?, fecha_ingreso=?, fecha_otorgamiento=?
                      WHERE id=?''',
                   (nueva_nomina, nuevo_nombre, nueva_area, nuevo_monto, diferencia, 
                    nuevo_autoriza, nuevo_id, nueva_fecha_ingreso, nueva_fecha_otorgamiento, p_id))
    
    cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual - ? WHERE sucursal = ?', (diferencia, sucursal))
    
    cursor.execute('INSERT INTO auditoria (fecha_hora, usuario, accion, empleado_nombre, detalle, motivo, sucursal) VALUES (?,?,?,?,?,?,?)',
                   (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user', 'SISTEMAS'), 'EDICIÓN DATOS', nuevo_nombre,
                    f"Cambio datos maestros. Monto modificado de ${monto_viejo} a ${nuevo_monto}. ID asignado: {nuevo_id}", "Edición desde Panel de Control", sucursal))
    
    conn.commit()
    conn.close()
    flash("Datos del préstamo actualizados y ID generado automáticamente.", "success")
    return redirect(request.referrer or '/')

# ------------------------------------------------------------
# 4. ABONOS Y AUDITORÍA
# ------------------------------------------------------------
@app.route('/registrar_abono', methods=['POST'])
def registrar_abono():
    if 'user' not in session:
        return redirect('/login')
    
    semana = request.form.get('semana_act')
    if not semana:
        semana = datetime.now().isocalendar()[1]
        
    ids_empleados = request.form.getlist('seleccionar_empleado')
    
    sucursal_actual = request.form.get("sucursal_act", "TODAS")
    if not sucursal_actual:
        sucursal_actual = "TODAS"
    
    if not ids_empleados:
        flash("Debes seleccionar empleados para procesar abonos.", "danger")
        return redirect(f'/semana/{semana}/sucursal/{sucursal_actual}')

    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    
    for emp_id in ids_empleados:
        monto_str = request.form.get(f'monto_abono_{emp_id}')
        if monto_str and float(monto_str) > 0:
            monto = int(float(monto_str))
            cursor.execute('INSERT INTO movimientos (prestamo_id, fecha, semana, tipo, monto) VALUES (?,?,?,?,?)',
                          (emp_id, fecha_hoy, semana, 'ABONO', monto))
            cursor.execute('UPDATE prestamos SET saldo_pendiente = round(saldo_pendiente - ?, 2) WHERE id = ?', (monto, emp_id))
            cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual + ? WHERE sucursal = (SELECT sucursal FROM prestamos WHERE id = ?)',
                          (monto, emp_id))
            suc_res = cursor.execute('SELECT sucursal FROM prestamos WHERE id = ?', (emp_id,)).fetchone()
            if suc_res:
                sucursal_actual = suc_res[0]
            
    conn.commit()
    conn.close()
    flash("Abonos procesados correctamente.", "success")
    return redirect(f'/semana/{semana}/sucursal/{sucursal_actual}')

@app.route('/eliminar_empleado/<int:id>')
def eliminar_empleado(id):
    if not es_admin():
        return redirect('/')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.execute('DELETE FROM movimientos WHERE prestamo_id = ?', (id,))
    conn.execute('DELETE FROM prestamos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash("Trabajador eliminado del sistema.", "success")
    return redirect(request.referrer or '/')

@app.route('/auditar_abonos/<int:prestamo_id>')
def auditar_abonos(prestamo_id):
    if 'user' not in session:
        return redirect('/login')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    emp = conn.execute('SELECT * FROM prestamos WHERE id = ?', (prestamo_id,)).fetchone()
    
    if not es_admin() and emp['sucursal'] != session['sucursal']:
        conn.close()
        flash("Acceso denegado.", "danger")
        return redirect('/')
        
    movs = conn.execute('SELECT * FROM movimientos WHERE prestamo_id = ? ORDER BY semana DESC, id DESC', (prestamo_id,)).fetchall()
    conn.close()
    return render_template('editar_abonos.html', emp=emp, movimientos=movs, admin=es_admin())

@app.route('/actualizar_movimiento', methods=['POST'])
def actualizar_movimiento():
    if not es_admin():
        return redirect('/')
    mov_id = request.form['mov_id']
    prestamo_id = request.form['prestamo_id']
    nuevo_monto = int(float(request.form['nuevo_monto']))
    nueva_semana = request.form['nueva_semana']
    motivo = request.form['motivo'] 

    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    emp = cursor.execute('SELECT empleado, sucursal FROM prestamos WHERE id = ?', (prestamo_id,)).fetchone()
    mov_viejo = cursor.execute('SELECT monto FROM movimientos WHERE id = ?', (mov_id,)).fetchone()
    monto_viejo = mov_viejo[0] if mov_viejo else 0
    diferencia = nuevo_monto - monto_viejo

    cursor.execute('UPDATE movimientos SET monto = ?, semana = ? WHERE id = ?', (nuevo_monto, nueva_semana, mov_id))
    cursor.execute('UPDATE prestamos SET saldo_pendiente = round(saldo_pendiente - ?, 2) WHERE id = ?', (diferencia, prestamo_id))
    cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual + ? WHERE sucursal = (SELECT sucursal FROM prestamos WHERE id = ?)', (diferencia, prestamo_id))
    
    cursor.execute('INSERT INTO auditoria (fecha_hora, usuario, accion, empleado_nombre, detalle, motivo, sucursal) VALUES (?,?,?,?,?,?,?)',
                   (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user', 'SISTEMAS'), 'EDICIÓN', emp[0],
                    f"Monto modificado: De ${monto_viejo} a ${nuevo_monto}", motivo, emp[1]))

    conn.commit()
    conn.close()
    flash("Movimiento corregido y documentado en bitácora.", "success")
    return redirect(f'/auditar_abonos/{prestamo_id}')

@app.route('/borrar_movimiento', methods=['POST'])
def borrar_movimiento():
    if not es_admin():
        return redirect('/')
    mov_id = request.form['mov_id']
    prestamo_id = request.form['prestamo_id']
    motivo = request.form['motivo'] 

    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    emp = cursor.execute('SELECT empleado, sucursal FROM prestamos WHERE id = ?', (prestamo_id,)).fetchone()
    mov_viejo = cursor.execute('SELECT monto, semana FROM movimientos WHERE id = ?', (mov_id,)).fetchone()
    
    if mov_viejo:
        monto_viejo = mov_viejo[0]
        semana = mov_viejo[1]
        cursor.execute('UPDATE prestamos SET saldo_pendiente = round(saldo_pendiente + ?, 2) WHERE id = ?', (monto_viejo, prestamo_id))
        cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual - ? WHERE sucursal = (SELECT sucursal FROM prestamos WHERE id = ?)', (monto_viejo, prestamo_id))
        cursor.execute('DELETE FROM movimientos WHERE id = ?', (mov_id,))
        
        cursor.execute('INSERT INTO auditoria (fecha_hora, usuario, accion, empleado_nombre, detalle, motivo, sucursal) VALUES (?,?,?,?,?,?,?)',
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user', 'SISTEMAS'), 'ELIMINACIÓN', emp[0],
                        f"Se eliminó abono de ${monto_viejo} (Semana {semana})", motivo, emp[1]))

    conn.commit()
    conn.close()
    flash("Abono eliminado y documentado en bitácora.", "warning")
    return redirect(f'/auditar_abonos/{prestamo_id}')

# ------------------------------------------------------------
# 5. REPORTES
# ------------------------------------------------------------
@app.route('/reporte/avanzado', methods=['POST'])
def r_avanzado():
    if 'user' not in session:
        return redirect('/login')
    
    tipo_reporte = request.form.get('tipo_reporte')
    f_inicio = request.form.get('fecha_inicio')
    f_fin = request.form.get('fecha_fin')
    suc = request.form.get('sucursal') if es_admin() else session['sucursal']
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    query = "SELECT *, (SELECT IFNULL(SUM(monto), 0) FROM movimientos WHERE prestamo_id = p.id AND tipo = 'ABONO') as total_abonado FROM prestamos p WHERE fecha_otorgamiento BETWEEN ? AND ?"
    params = [f_inicio, f_fin]

    if suc != 'TODAS':
        query += " AND sucursal = ?"
        params.append(suc)

    if tipo_reporte == 'solo_deudores':
        query += " AND saldo_pendiente > 0"
        titulo = "Reporte de Empleados con Adeudo"
    elif tipo_reporte == 'multi_prestamo':
        query += " AND nomina IN (SELECT nomina FROM prestamos GROUP BY nomina HAVING COUNT(*) > 1)"
        titulo = "Reporte de Empleados con Más de 1 Préstamo"
    else:
        titulo = "Reporte Histórico Filtrado"

    prestamos = conn.execute(query + " ORDER BY sucursal ASC, empleado ASC", params).fetchall()
    
    tot_inicial = sum(p['monto_inicial'] for p in prestamos)
    tot_pendiente = sum(p['saldo_pendiente'] for p in prestamos)
    tot_recuperado = sum(p['total_abonado'] for p in prestamos)
    conn.close()
    
    return render_template('reporte_especial.html',
                          prestamos=prestamos,
                          sucursal=suc,
                          tot_inicial=tot_inicial,
                          tot_pendiente=tot_pendiente,
                          tot_recuperado=tot_recuperado,
                          titulo=titulo,
                          f_inicio=f_inicio,
                          f_fin=f_fin)

@app.route('/reportes')
def reportes():
    if 'user' not in session:
        return redirect('/login')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    if es_admin():
        empleados = conn.execute('''
            SELECT MIN(id) as id, nomina, empleado, sucursal, 
                   MIN(id_empleado) as id_empleado
            FROM prestamos 
            GROUP BY nomina, empleado
            ORDER BY sucursal ASC, empleado ASC
        ''').fetchall()
    else:
        empleados = conn.execute('''
            SELECT MIN(id) as id, nomina, empleado, sucursal,
                   MIN(id_empleado) as id_empleado
            FROM prestamos 
            WHERE sucursal = ?
            GROUP BY nomina, empleado
            ORDER BY empleado ASC
        ''', (session['sucursal'],)).fetchall()
    
    conn.close()
    return render_template('reportes.html', empleados=empleados, admin=es_admin(), suc_act=session.get('sucursal'))

@app.route('/reporte/historial_empresa', methods=['POST'])
def r_historial_empresa():
    if 'user' not in session:
        return redirect('/login')
    suc = request.form.get('sucursal') if es_admin() else session['sucursal']
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    query = '''
        SELECT p.*, 
        (SELECT IFNULL(SUM(monto), 0) FROM movimientos WHERE prestamo_id = p.id AND tipo = 'ABONO') as total_abonado
        FROM prestamos p
        WHERE p.saldo_pendiente > 0
    '''
    
    if suc == 'TODAS' and es_admin():
        prestamos = conn.execute(query + ' ORDER BY sucursal ASC, empleado ASC').fetchall()
        tot_inicial = conn.execute('SELECT SUM(monto_inicial) FROM prestamos WHERE saldo_pendiente > 0').fetchone()[0] or 0
        tot_pendiente = conn.execute('SELECT SUM(saldo_pendiente) FROM prestamos WHERE saldo_pendiente > 0').fetchone()[0] or 0
    else:
        prestamos = conn.execute(query + ' AND p.sucursal = ? ORDER BY empleado ASC', (suc,)).fetchall()
        tot_inicial = conn.execute('SELECT SUM(monto_inicial) FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (suc,)).fetchone()[0] or 0
        tot_pendiente = conn.execute('SELECT SUM(saldo_pendiente) FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (suc,)).fetchone()[0] or 0
    
    tot_recuperado = tot_inicial - tot_pendiente
    
    conn.close()
    
    return render_template('reporte_historial.html',
                          prestamos=prestamos,
                          sucursal=suc,
                          tot_inicial=tot_inicial,
                          tot_pendiente=tot_pendiente,
                          tot_recuperado=tot_recuperado)

@app.route('/reporte/kardex_empleado', methods=['POST'])
def r_kardex_emp():
    emp_id = request.form.get('empleado_id')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    # Obtener el préstamo seleccionado para conocer la nómina y la sucursal
    prestamo_seleccionado = conn.execute('SELECT nomina, sucursal FROM prestamos WHERE id = ?', (emp_id,)).fetchone()
    if not prestamo_seleccionado:
        conn.close()
        flash("Empleado no encontrado", "danger")
        return redirect('/reportes')
    
    nomina = prestamo_seleccionado['nomina']
    sucursal_origen = prestamo_seleccionado['sucursal']
    
    # Siempre filtrar por la sucursal del préstamo seleccionado
    query = '''
        SELECT * FROM prestamos 
        WHERE nomina = ? AND sucursal = ?
    '''
    params = [nomina, sucursal_origen]
    
    todos_prestamos = conn.execute(query, params).fetchall()
    
    if not todos_prestamos:
        conn.close()
        flash("No se encontraron préstamos para este empleado en esta sucursal.", "danger")
        return redirect('/reportes')
    
    # Tomar el primer préstamo para datos base del empleado
    emp = todos_prestamos[0]
    
    prestamos_con_movimientos = []
    for prestamo in todos_prestamos:
        movimientos = conn.execute('''
            SELECT * FROM movimientos 
            WHERE prestamo_id = ? 
            ORDER BY semana DESC
        ''', (prestamo['id'],)).fetchall()
        
        total_abonado = sum(m['monto'] for m in movimientos if m['tipo'] == 'ABONO')
        
        prestamos_con_movimientos.append({
            'prestamo': prestamo,
            'movimientos': movimientos,
            'total_abonado': total_abonado
        })
    
    conn.close()
    
    return render_template('kardex_detalle.html', 
                          emp=emp,
                          prestamos=prestamos_con_movimientos)

@app.route('/reporte/kardex_semana', methods=['POST'])
def r_kardex_sem():
    try:
        sem = int(request.form.get('semana'))
    except:
        sem = int(datetime.now().isocalendar()[1])
        
    suc = request.form.get('sucursal') if es_admin() else session['sucursal']
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    if suc == 'TODAS' and es_admin():
        prestamos = conn.execute('''
            SELECT p.* FROM prestamos p 
            WHERE CAST(IFNULL(p.semana_otorgada, strftime('%W', p.fecha_otorgamiento)) AS INTEGER) = ? 
            OR EXISTS (SELECT 1 FROM movimientos m WHERE m.prestamo_id = p.id AND CAST(m.semana AS INTEGER) = ?) 
            ORDER BY p.sucursal ASC, p.empleado ASC''', (sem, sem)).fetchall()
    else:
        prestamos = conn.execute('''
            SELECT p.* FROM prestamos p 
            WHERE (CAST(IFNULL(p.semana_otorgada, strftime('%W', p.fecha_otorgamiento)) AS INTEGER) = ? 
            OR EXISTS (SELECT 1 FROM movimientos m WHERE m.prestamo_id = p.id AND CAST(m.semana AS INTEGER) = ?)) 
            AND p.sucursal = ? 
            ORDER BY p.empleado ASC''', (sem, sem, suc)).fetchall()
    
    data = []
    for p in prestamos:
        movs = conn.execute('SELECT * FROM movimientos WHERE prestamo_id = ? ORDER BY semana DESC', (p['id'],)).fetchall()
        data.append({'emp': p, 'movimientos': movs})
        
    conn.close()
    return render_template('kardex_semana.html', data=data, semana=sem, sucursal=suc)

@app.route('/reporte/auditoria', methods=['POST', 'GET'])
def r_auditoria():
    if not es_admin():
        return redirect('/')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    registros = conn.execute('SELECT * FROM auditoria ORDER BY id DESC').fetchall()
    conn.close()
    return render_template('reporte_auditoria.html', registros=registros)

# ------------------------------------------------------------
# REPORTE DE ARQUEOS (CUADRE DE CAJA)
# ------------------------------------------------------------
@app.route('/reporte_arqueo', methods=['POST'])
def reporte_arqueo():
    if 'user' not in session:
        return redirect('/login')
    
    sucursal = request.form.get('sucursal')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    if sucursal == 'TODAS' and es_admin():
        registros = conn.execute('SELECT * FROM arqueos ORDER BY fecha DESC').fetchall()
    else:
        if not es_admin():
            sucursal = session.get('sucursal')
        registros = conn.execute('SELECT * FROM arqueos WHERE sucursal = ? ORDER BY fecha DESC', (sucursal,)).fetchall()
    
    fecha_generacion = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.close()
    
    return render_template('reporte_arqueo.html', 
                           registros=registros, 
                           sucursal=sucursal, 
                           fecha_generacion=fecha_generacion)

# ------------------------------------------------------------
# 6. GESTIÓN DE USUARIOS
# ------------------------------------------------------------
@app.route('/editar_usuario', methods=['POST'])
def editar_usuario():
    if not es_admin():
        return redirect('/')
    u_id, u_name, u_suc, new_pass = request.form['id'], request.form['username'], request.form['sucursal'], request.form['password']
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    if new_pass:
        pw = generate_password_hash(new_pass)
        conn.execute('UPDATE usuarios SET username=?, password=?, sucursal=? WHERE id=?', (u_name, pw, u_suc, u_id))
    else:
        conn.execute('UPDATE usuarios SET username=?, sucursal=? WHERE id=?', (u_name, u_suc, u_id))
    conn.commit()
    conn.close()
    flash("Usuario actualizado.", "success")
    return redirect('/')

@app.route('/eliminar_usuario/<int:id>')
def eliminar_usuario(id):
    if not es_admin():
        return redirect('/')
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.execute('DELETE FROM usuarios WHERE id=?', (id,))
    conn.commit()
    conn.close()
    flash("Usuario eliminado.", "success")
    return redirect('/')

@app.route('/crear_usuario', methods=['POST'])
def crear_usuario():
    if not es_admin():
        return redirect('/')
    u, p, s = request.form['username'], generate_password_hash(request.form['password']), request.form['sucursal']
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    try:
        conn.execute('INSERT INTO usuarios (username, password, rol, sucursal) VALUES (?,?,?,?)', (u, p, 'SUCURSAL', s))
        conn.commit()
        flash("Usuario creado.", "success")
    except:
        flash("Error: El usuario ya existe.", "danger")
    conn.close()
    return redirect('/')

# ------------------------------------------------------------
# 7. TRASPASO ANUAL
# ------------------------------------------------------------
@app.route('/cambiar_anio', methods=['POST'])
def cambiar_anio():
    if not es_admin():
        return redirect('/')
    nuevo_anio = request.form['nuevo_anio']
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    prestamos_deuda = cursor.execute('SELECT * FROM prestamos WHERE saldo_pendiente > 0').fetchall()
    for p in prestamos_deuda:
        nuevo_id = generar_id_empleado(p['sucursal'], p['nomina'], conn)
        cursor.execute('''INSERT INTO prestamos (nomina, empleado, area, monto_inicial, saldo_pendiente, fecha_otorgamiento, sucursal, semana_otorgada, id_empleado, fecha_ingreso) 
                          VALUES (?,?,?,?,?,?,?,?,?,?)''',
                       (p['nomina'], p['empleado'], p['area'], p['saldo_pendiente'], p['saldo_pendiente'],
                        f"{nuevo_anio}-01-01", p['sucursal'], 1, nuevo_id, p['fecha_ingreso']))
        cursor.execute('UPDATE prestamos SET saldo_pendiente = 0 WHERE id = ?', (p['id'],))
    conn.commit()
    conn.close()
    flash("Traspaso completado.", "success")
    return redirect('/')

# ------------------------------------------------------------
# 8. IMPORTACIÓN Y NUEVO PRÉSTAMO
# ------------------------------------------------------------
@app.route('/importar_excel', methods=['POST'])
def importar_excel():
    if not es_admin():
        return redirect('/')
    file = request.files['archivo_excel']
    suc_dest = request.form['sucursal_importacion']
    if file:
        try:
            df = pd.read_csv(file, encoding='latin1') if file.filename.endswith('.csv') else pd.read_excel(file)
            df.columns = [str(c).upper().strip() for c in df.columns]
            df = df.loc[:, ~df.columns.duplicated()]
            
            conn = sqlite3.connect(DB_NAME, timeout=10.0)
            cursor = conn.cursor()
            for _, row in df.iterrows():
                nombre = row.get('NOMBRE', row.get('EMPLEADO', None))
                if pd.isna(nombre) or str(nombre).strip() == '':
                    continue
                
                nomina = str(row.get('N.NO', row.get('NOMINA', 'S/N'))).replace('.0', '')
                area = str(row.get('AREA', 'GENERAL'))
                
                monto_raw = str(row.get('MONTO', row.get('PRESTAMO', 0)))
                monto_limpio = re.sub(r'[^\d.]', '', monto_raw)
                monto = float(monto_limpio) if monto_limpio else 0.0
                
                fecha_raw = row.get('FECHA')
                if pd.isna(fecha_raw) or str(fecha_raw).strip() == '':
                    fecha = datetime.now().strftime('%Y-%m-%d')
                else:
                    fecha_str = str(fecha_raw).strip()
                    try:
                        fecha = datetime.strptime(fecha_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                    except:
                        fecha = fecha_str 

                sem_raw = row.get('SEM', row.get('SEMANA', None))
                if pd.notna(sem_raw) and str(sem_raw).strip() != '':
                    semana_otorgada = int(float(sem_raw))
                else:
                    try:
                        semana_otorgada = datetime.strptime(fecha, '%Y-%m-%d').isocalendar()[1]
                    except:
                        semana_otorgada = datetime.now().isocalendar()[1]
                
                id_empleado = generar_id_empleado(suc_dest, nomina, conn)
                fecha_ingreso = None
                if 'F.ALTA' in row:
                    fecha_ingreso = row['F.ALTA']
                    if pd.notna(fecha_ingreso):
                        try:
                            fecha_ingreso = datetime.strptime(str(fecha_ingreso), '%d/%m/%Y').strftime('%Y-%m-%d')
                        except:
                            fecha_ingreso = str(fecha_ingreso)
                        
                cursor.execute('''INSERT INTO prestamos 
                                  (nomina, empleado, area, monto_inicial, saldo_pendiente, fecha_otorgamiento, sucursal, semana_otorgada, id_empleado, fecha_ingreso) 
                                  VALUES (?,?,?,?,?,?,?,?,?,?)''',
                               (nomina, nombre, area, monto, monto, fecha, suc_dest, semana_otorgada, id_empleado, fecha_ingreso))
                cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual - ? WHERE sucursal = ?', (monto, suc_dest))
                
            conn.commit()
            conn.close()
            flash("Importación exitosa. Préstamos registrados.", "success")
        except Exception as e:
            print(f"Error importando archivo: {e}")
            flash("Error al leer el archivo. Verifica el formato.", "danger")
            
    return redirect('/')

@app.route('/nuevo_prestamo', methods=['POST'])
def nuevo_prestamo():
    if 'user' not in session:
        return redirect('/login')
    n = request.form['nomina']
    nom = request.form['nombre']
    p = request.form.get('puesto', '')
    a = request.form['area']
    m = int(float(request.form['monto']))
    f = request.form['fecha_otorgamiento']
    s = request.form['sucursal'] if es_admin() else session['sucursal']
    fecha_ingreso = request.form.get('fecha_ingreso', '')
    
    autoriza = request.form.get('autoriza', '')
    motivo = request.form.get('motivo', '')
    semana_otorgada = datetime.strptime(f, '%Y-%m-%d').isocalendar()[1]
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    # Verificar si el empleado tiene un préstamo vigente en la MISMA sucursal
    existente = cursor.execute('SELECT saldo_pendiente FROM prestamos WHERE nomina = ? AND sucursal = ? AND saldo_pendiente > 0', (n, s)).fetchone()
    if existente and existente[0] > 0 and str(motivo).strip() == '':
        flash(f"El empleado {nom} ya tiene un préstamo vigente de ${existente[0]:,.2f} en esta sucursal. Debes ingresar un MOTIVO obligatorio para agregarle otro.", "danger")
        conn.close()
        return redirect('/')
    
    id_empleado = generar_id_empleado(s, n, conn)
    
    cursor.execute('''INSERT INTO prestamos 
                      (nomina, empleado, area, puesto, monto_inicial, saldo_pendiente, fecha_otorgamiento, sucursal, semana_otorgada, autoriza, motivo_adicional, id_empleado, fecha_ingreso) 
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                   (n, nom, a, p, m, m, f, s, semana_otorgada, autoriza, motivo, id_empleado, fecha_ingreso))
    cursor.execute('UPDATE cajas SET saldo_actual = saldo_actual - ? WHERE sucursal = ?', (m, s))
    conn.commit()
    conn.close()
    flash(f"Préstamo registrado exitosamente con ID: {id_empleado}", "success")
    return redirect('/')

# ------------------------------------------------------------
# 9. API PARA AUTOCOMPLETADO (DESDE EXCEL DE DEPARTAMENTOS)
# ------------------------------------------------------------
@app.route('/api/empleados_excel/<sucursal>')
def api_empleados_excel(sucursal):
    if 'user' not in session:
        return jsonify([])
    
    if sucursal == 'TODAS':
        return jsonify([])
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(base_dir, 'empleados')
    
    archivo_encontrado = None
    for nombre in os.listdir(folder_path):
        if nombre.upper().startswith(f"DEPARTAMENTO {sucursal.upper()}") or nombre.upper().startswith(f"DEPARTAMENTO_{sucursal.upper()}"):
            archivo_encontrado = nombre
            break
    if not archivo_encontrado:
        for nombre in os.listdir(folder_path):
            if nombre.upper().startswith(sucursal.upper()):
                archivo_encontrado = nombre
                break
    
    if not archivo_encontrado:
        return jsonify([])
    
    filepath = os.path.join(folder_path, archivo_encontrado)
    
    try:
        df_headers = pd.read_excel(filepath, nrows=1, header=None)
        num_cols = len(df_headers.columns)
        fecha_col_idx = 5 if num_cols >= 6 else 3
        puesto_col_idx = 4 if num_cols >= 6 else 2
        
        df = pd.read_excel(filepath, header=None, dtype=str)
        empleados = []
        departamento_actual = ""
        
        for idx, row in df.iterrows():
            primera_col = str(row[0]) if pd.notna(row[0]) else ""
            if "Departamento:" in primera_col:
                if "/" in primera_col:
                    partes = primera_col.split("/")
                    if len(partes) >= 2:
                        departamento_actual = partes[1].strip()
                    else:
                        departamento_actual = primera_col.replace("Departamento:", "").strip()
                else:
                    departamento_actual = primera_col.replace("Departamento:", "").strip()
                continue
            
            clave = str(row[0]).strip() if pd.notna(row[0]) else ""
            nombre = str(row[1]).strip() if pd.notna(row[1]) else ""
            
            if clave == "" or nombre == "" or clave == "nan" or nombre == "nan" or clave == "Clave":
                continue
            
            puesto = ""
            if len(row) > puesto_col_idx and pd.notna(row[puesto_col_idx]):
                puesto = str(row[puesto_col_idx]).strip()
            
            fecha_alta_raw = None
            if len(row) > fecha_col_idx and pd.notna(row[fecha_col_idx]):
                fecha_alta_raw = row[fecha_col_idx]
            
            fecha_alta = ""
            if fecha_alta_raw:
                fecha_str = str(fecha_alta_raw).strip()
                try:
                    fecha_str = fecha_str.replace('-', '/')
                    fecha_alta = datetime.strptime(fecha_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                except:
                    fecha_alta = fecha_str
            
            empleados.append({
                'nomina': clave,
                'nombre': nombre,
                'puesto': puesto,
                'departamento': departamento_actual,
                'fecha_ingreso': fecha_alta
            })
        
        return jsonify(empleados)
    except Exception as e:
        print(f"Error leyendo excel {sucursal}: {e}")
        return jsonify([])
    
# ------------------------------------------------------------
# 10. GRÁFICAS POR SEDE
# ------------------------------------------------------------
@app.route('/graficas')
def graficas():
    if 'user' not in session:
        return redirect('/login')
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    if es_admin():
        sucursales = ['TACUBA', 'BOMBILLA', 'RESTORANES', 'NAPOLES', 'BRIGAR', 'BGARI']
        titulo = "Gráfica Global de Rendimiento (Préstamos Activos)"
    else:
        sucursales = [session['sucursal']]
        titulo = f"Gráfica de Rendimiento Activo - {session['sucursal']}"
        
    labels, prestado, recuperado, pendiente = [], [], [], []
    for suc in sucursales:
        stats = conn.execute('SELECT SUM(monto_inicial) as prest, SUM(saldo_pendiente) as pend FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (suc,)).fetchone()
        p = stats['prest'] or 0
        pd = stats['pend'] or 0
        r = p - pd
        labels.append(suc)
        prestado.append(p)
        recuperado.append(r)
        pendiente.append(pd)
    conn.close()
    
    return render_template('graficas.html',
                          labels=labels,
                          prestado=prestado,
                          recuperado=recuperado,
                          pendiente=pendiente,
                          admin=es_admin(),
                          titulo=titulo)

# ------------------------------------------------------------
# 11. RESUMEN POR EMPRESA Y ARQUEOS
# ------------------------------------------------------------
@app.route('/resumen_empresa')
def resumen_empresa_default():
    if not es_admin():
        return redirect(f'/resumen_empresa/{session.get("sucursal")}')
    return redirect('/resumen_empresa/TACUBA')

@app.route('/resumen_empresa/<sucursal>')
def resumen_empresa(sucursal):
    if not es_admin() and session.get('sucursal') != sucursal:
        return redirect(f'/resumen_empresa/{session.get("sucursal")}')
    
    sucursales_validas = ['TACUBA', 'BOMBILLA', 'RESTORANES', 'NAPOLES', 'BRIGAR', 'BGARI']
    if sucursal not in sucursales_validas:
        sucursal = 'TACUBA'
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    caja = conn.execute('SELECT saldo_inicial, saldo_actual FROM cajas WHERE sucursal = ?', (sucursal,)).fetchone()
    fondo_inicial = int(caja['saldo_inicial']) if caja else 0
    fondo_actual = int(caja['saldo_actual']) if caja else 0
    
    ultimo_arqueo = conn.execute('SELECT * FROM arqueos WHERE sucursal = ? ORDER BY fecha DESC LIMIT 1', (sucursal,)).fetchone()
    
    hoy = datetime.now()
    semana_actual = hoy.isocalendar()[1]
    arqueo_pendiente = False
    
    if not (es_admin() and sucursal != session.get('sucursal')):
        arqueo_hecho = conn.execute('SELECT id FROM arqueos WHERE sucursal = ? AND semana = ? AND strftime("%Y", fecha) = ?', 
                                    (sucursal, semana_actual, str(hoy.year))).fetchone()
        if not arqueo_hecho:
            arqueo_pendiente = True
            
    historial_arqueos = conn.execute('SELECT * FROM arqueos WHERE sucursal = ? ORDER BY id DESC', (sucursal,)).fetchall()
    
    # Parsear detalle y agregar detalle_json
    historial_arqueos_parsed = []
    for a in historial_arqueos:
        a_dict = dict(a)
        try:
            if a_dict.get('detalle'):
                a_dict['detalle'] = json.loads(a_dict['detalle'])
            else:
                a_dict['detalle'] = {}
        except:
            a_dict['detalle'] = {}
        # Generar JSON string para la vista (con ensure_ascii=False para caracteres especiales)
        a_dict['detalle_json'] = json.dumps(a_dict['detalle'], ensure_ascii=False)
        historial_arqueos_parsed.append(a_dict)
    
    prestamos_activos = conn.execute('''
        SELECT id, id_empleado, nomina, empleado, area, puesto, monto_inicial, saldo_pendiente, fecha_ingreso,
               (monto_inicial - saldo_pendiente) as recuperado
        FROM prestamos 
        WHERE sucursal = ? AND saldo_pendiente > 0
        ORDER BY empleado ASC
    ''', (sucursal,)).fetchall()
    
    total_prestado = conn.execute('SELECT SUM(monto_inicial) FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (sucursal,)).fetchone()[0] or 0
    total_adeudo = conn.execute('SELECT SUM(saldo_pendiente) FROM prestamos WHERE sucursal = ? AND saldo_pendiente > 0', (sucursal,)).fetchone()[0] or 0
    total_recaudado = total_prestado - total_adeudo
    
    cajas = conn.execute('SELECT * FROM cajas').fetchall()
    usuarios_lista = conn.execute('SELECT * FROM usuarios WHERE username != "SISTEMAS"').fetchall()
    res_anio = conn.execute('SELECT fecha_otorgamiento FROM prestamos ORDER BY id DESC LIMIT 1').fetchone()
    anio_act = str(hoy.year)
    if res_anio and res_anio[0]:
        match = re.search(r'\d{4}', str(res_anio[0]))
        if match:
            anio_act = match.group()
    
    conn.close()
    
    return render_template('resumen_empresa.html',
                          sucursal=sucursal, sucursales=sucursales_validas,
                          fondo_inicial=fondo_inicial, fondo_actual=fondo_actual,
                          total_prestado=total_prestado, total_recaudado=total_recaudado,
                          total_adeudo=total_adeudo, prestamos=prestamos_activos,
                          anio_act=anio_act, admin=es_admin(), cajas=cajas,
                          usuarios=usuarios_lista, datetime=datetime,
                          arqueo_pendiente=arqueo_pendiente, historial_arqueos=historial_arqueos_parsed,
                          ultimo_arqueo=ultimo_arqueo)

# ================= FUNCIONES AUXILIARES PARA CONVERSIÓN SEGURA =================
def safe_int(val):
    try:
        if val is None or str(val).strip() == '':
            return 0
        return int(float(val))
    except:
        return 0

def safe_float(val):
    try:
        if val is None or str(val).strip() == '':
            return 0.0
        return float(val)
    except:
        return 0.0

@app.route('/guardar_arqueo', methods=['POST'])
def guardar_arqueo():
    if 'user' not in session:
        return redirect('/login')
    
    sucursal = request.form.get('sucursal')
    fondo_sistema = safe_float(request.form.get('fondo_sistema'))
    efectivo_real = safe_float(request.form.get('efectivo_real'))
    observaciones = request.form.get('observaciones', '')
    
    # Capturar detalle de billetes y monedas con conversión segura
    detalle = {
        '1000': safe_int(request.form.get('billete_1000')),
        '500': safe_int(request.form.get('billete_500')),
        '200': safe_int(request.form.get('billete_200')),
        '100': safe_int(request.form.get('billete_100')),
        '50': safe_int(request.form.get('billete_50')),
        '20': safe_int(request.form.get('billete_20')),
        'monedas': safe_float(request.form.get('monedas'))
    }
    detalle_json = json.dumps(detalle)
    
    diferencia = efectivo_real - fondo_sistema
    hoy = datetime.now()
    semana = hoy.isocalendar()[1]
    fecha_str = hoy.strftime('%Y-%m-%d %H:%M:%S')
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.execute('INSERT INTO arqueos (sucursal, fecha, semana, fondo_sistema, efectivo_real, diferencia, observaciones, usuario, detalle) VALUES (?,?,?,?,?,?,?,?,?)',
                 (sucursal, fecha_str, semana, fondo_sistema, efectivo_real, diferencia, observaciones, session['user'], detalle_json))
    conn.commit()
    conn.close()
    
    flash("Arqueo registrado exitosamente. Gracias por cumplir con el proceso.", "success")
    return redirect(f'/resumen_empresa/{sucursal}')

# =============== EDICIÓN Y ELIMINACIÓN DE ARQUEOS ===============
@app.route('/editar_arqueo/<int:id>', methods=['POST'])
def editar_arqueo(id):
    if not es_admin():
        return redirect('/')
    
    nuevo_efectivo = safe_float(request.form.get('efectivo_real'))
    nuevas_obs = request.form.get('observaciones', '')
    
    detalle = {
        '1000': safe_int(request.form.get('billete_1000')),
        '500': safe_int(request.form.get('billete_500')),
        '200': safe_int(request.form.get('billete_200')),
        '100': safe_int(request.form.get('billete_100')),
        '50': safe_int(request.form.get('billete_50')),
        '20': safe_int(request.form.get('billete_20')),
        'monedas': safe_float(request.form.get('monedas'))
    }
    detalle_json = json.dumps(detalle)
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    arqueo = cursor.execute('SELECT * FROM arqueos WHERE id = ?', (id,)).fetchone()
    if not arqueo:
        conn.close()
        flash("Arqueo no encontrado.", "danger")
        return redirect(request.referrer or '/')
    
    fondo_sistema = arqueo[4]
    nueva_diferencia = nuevo_efectivo - fondo_sistema
    
    cursor.execute('''UPDATE arqueos 
                      SET efectivo_real = ?, diferencia = ?, observaciones = ?, detalle = ?
                      WHERE id = ?''',
                   (nuevo_efectivo, nueva_diferencia, nuevas_obs, detalle_json, id))
    conn.commit()
    conn.close()
    
    flash(f"Arqueo #{id} actualizado correctamente (incluye detalle de billetes).", "success")
    return redirect(request.referrer or f'/resumen_empresa/{arqueo[1]}')

@app.route('/eliminar_arqueo/<int:id>', methods=['POST'])
def eliminar_arqueo(id):
    if not es_admin():
        return redirect('/')
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    cursor = conn.cursor()
    
    arqueo = cursor.execute('SELECT * FROM arqueos WHERE id = ?', (id,)).fetchone()
    if not arqueo:
        conn.close()
        flash("Arqueo no encontrado.", "danger")
        return redirect(request.referrer or '/')
    
    sucursal = arqueo[1]
    cursor.execute('DELETE FROM arqueos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    flash(f"Arqueo #{id} eliminado permanentemente.", "warning")
    return redirect(request.referrer or f'/resumen_empresa/{sucursal}')

# ================= EXPORTAR KARDEX DE ARQUEOS =================
def crear_excel_profesional(dataframe, nombre_empresa, titulo, output):
    """Crea un Excel profesional con estilos a partir de un DataFrame"""
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False, sheet_name='Kardex Arqueos')
        wb = writer.book
        ws = wb['Kardex Arqueos']
        
        # Insertar filas de título
        ws.insert_rows(1, 4)
        
        # Título del reporte
        title_cell = ws['A1']
        title_cell.value = titulo
        title_cell.font = Font(size=16, bold=True, color='1A1D20')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:P1')
        
        # Nombre de la empresa
        empresa_cell = ws['A2']
        empresa_cell.value = nombre_empresa
        empresa_cell.font = Font(size=12, bold=True, color='333333')
        empresa_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:P2')
        
        # Fecha de emisión
        fecha_cell = ws['A3']
        fecha_cell.value = f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        fecha_cell.font = Font(size=10, color='666666')
        fecha_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A3:P3')
        
        # Aplicar estilos profesionales a los encabezados (fila 5)
        header_fill = PatternFill(start_color='1A1D20', end_color='1A1D20', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[5]:  # Fila 5 es donde están los encabezados después de insertar 4 filas
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Colores alternados para filas
        green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):  # Desde fila 6 hasta el final
            for cell in row:
                if row[0].row % 2 == 0:
                    cell.fill = green_fill
                else:
                    cell.fill = white_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formato de moneda para columnas numéricas
        currency_fmt = '#,##0.00'
        col_mapping = {
            'Fondo Sistema': currency_fmt,
            'Efectivo Físico': currency_fmt,
            'Diferencia': currency_fmt
        }
        
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                col_name = ws.cell(row=5, column=cell.column).value
                if col_name in col_mapping:
                    cell.number_format = currency_fmt
        
        # Bordes finos
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[col_letter].width = adjusted_width

@app.route('/exportar_kardex_arqueo/<int:id>')
def exportar_kardex_arqueo(id):
    if not es_admin():
        return redirect('/')
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    arqueo = conn.execute('SELECT * FROM arqueos WHERE id = ?', (id,)).fetchone()
    conn.close()
    
    if not arqueo:
        flash("Arqueo no encontrado.", "danger")
        return redirect(request.referrer or '/')
    
    # Convertir a diccionario y parsear detalle
    a_dict = dict(arqueo)
    try:
        detalle = json.loads(a_dict.get('detalle', '{}'))
    except:
        detalle = {}
    
    # Construir DataFrame con una fila
    data = {
        'ID': [a_dict['id']],
        'Fecha': [a_dict['fecha']],
        'Semana': [a_dict['semana']],
        'Sucursal': [a_dict['sucursal']],
        'Fondo Sistema': [a_dict['fondo_sistema']],
        'Efectivo Físico': [a_dict['efectivo_real']],
        'Diferencia': [a_dict['diferencia']],
        'Usuario': [a_dict['usuario']],
        'Observaciones': [a_dict['observaciones'] or ''],
        'Billetes $1000': [detalle.get('1000', 0)],
        'Billetes $500': [detalle.get('500', 0)],
        'Billetes $200': [detalle.get('200', 0)],
        'Billetes $100': [detalle.get('100', 0)],
        'Billetes $50': [detalle.get('50', 0)],
        'Billetes $20': [detalle.get('20', 0)],
        'Monedas/Otros': [detalle.get('monedas', 0)]
    }
    
    df = pd.DataFrame(data)
    
    # Generar Excel profesional
    output = io.BytesIO()
    nombre_empresa = "EL CARDENAL"
    titulo = f"REPORTE DE ARQUEO - SUCURSAL {a_dict['sucursal']} - #{a_dict['id']}"
    crear_excel_profesional(df, nombre_empresa, titulo, output)
    output.seek(0)
    
    nombre_archivo = f"Kardex_Arqueo_{a_dict['id']}_{a_dict['fecha'][:10]}.xlsx"
    return send_file(output, download_name=nombre_archivo, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar_kardex_arqueos', methods=['POST'])
def exportar_kardex_arqueos():
    if not es_admin():
        return redirect('/')
    
    ids = request.form.getlist('ids')
    if not ids:
        flash("No seleccionaste ningún arqueo.", "warning")
        return redirect(request.referrer or '/')
    
    # Convertir a enteros
    ids = [int(i) for i in ids if i.isdigit()]
    
    if not ids:
        flash("IDs inválidos.", "danger")
        return redirect(request.referrer or '/')
    
    conn = sqlite3.connect(DB_NAME, timeout=10.0)
    conn.row_factory = sqlite3.Row
    placeholders = ','.join('?' * len(ids))
    arqueos = conn.execute(f'SELECT * FROM arqueos WHERE id IN ({placeholders})', ids).fetchall()
    conn.close()
    
    if not arqueos:
        flash("No se encontraron arqueos.", "danger")
        return redirect(request.referrer or '/')
    
    # Construir lista de diccionarios
    rows = []
    for a in arqueos:
        a_dict = dict(a)
        try:
            detalle = json.loads(a_dict.get('detalle', '{}'))
        except:
            detalle = {}
        rows.append({
            'ID': a_dict['id'],
            'Fecha': a_dict['fecha'],
            'Semana': a_dict['semana'],
            'Sucursal': a_dict['sucursal'],
            'Fondo Sistema': a_dict['fondo_sistema'],
            'Efectivo Físico': a_dict['efectivo_real'],
            'Diferencia': a_dict['diferencia'],
            'Usuario': a_dict['usuario'],
            'Observaciones': a_dict['observaciones'] or '',
            'Billetes $1000': detalle.get('1000', 0),
            'Billetes $500': detalle.get('500', 0),
            'Billetes $200': detalle.get('200', 0),
            'Billetes $100': detalle.get('100', 0),
            'Billetes $50': detalle.get('50', 0),
            'Billetes $20': detalle.get('20', 0),
            'Monedas/Otros': detalle.get('monedas', 0)
        })
    
    df = pd.DataFrame(rows)
    
    # Generar Excel profesional
    output = io.BytesIO()
    nombre_empresa = "EL CARDENAL"
    titulo = f"REPORTE DE ARQUEOS - {len(rows)} REGISTROS SELECCIONADOS"
    crear_excel_profesional(df, nombre_empresa, titulo, output)
    output.seek(0)
    
    nombre_archivo = f"Kardex_Arqueos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=nombre_archivo, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ------------------------------------------------------------
# 12. ACCESO
# ------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u, p = request.form['username'], request.form['password']
        conn = sqlite3.connect(DB_NAME, timeout=10.0)
        conn.row_factory = sqlite3.Row
        user = conn.execute('SELECT * FROM usuarios WHERE username=?', (u,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password'], p):
            session['user'] = user['username']
            session['rol'] = user['rol']
            session['sucursal'] = user['sucursal']
            session.permanent = True
            
            token = str(uuid.uuid4())
            session['session_token'] = token
            sesiones_activas[user['username']] = {
                'token': token,
                'ultima_actividad': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'sucursal': user['sucursal']
            }
            return redirect('/')
        else:
            flash("Usuario o contraseña incorrectos.", "danger")
    return render_template('login.html')

@app.route('/logout')
def logout():
    user = session.get('user')
    if user in sesiones_activas:
        del sesiones_activas[user]
    session.clear()
    return redirect('/login')

# ------------------------------------------------------------
# RUTA PARA EL MANUAL DE AYUDA SEGÚN ROL
# ------------------------------------------------------------
@app.route('/manual_ayuda')
def manual_ayuda():
    if 'user' not in session:
        return redirect('/login')
    
    if es_admin():
        return render_template('manual_el_cardenal.html')
    else:
        return render_template('manual_sucursal_profesional.html')

# ------------------------------------------------------------
# 13. VISTA DE SESIONES ACTIVAS
# ------------------------------------------------------------
@app.route('/ver_sesiones')
def ver_sesiones():
    if not es_admin():
        return redirect('/')
    
    ahora = datetime.now()
    usuarios_inactivos = []
    for u, datos in sesiones_activas.items():
        try:
            ultima = datetime.strptime(datos['ultima_actividad'], '%Y-%m-%d %H:%M:%S')
            if (ahora - ultima).total_seconds() > 43200:
                usuarios_inactivos.append(u)
        except: pass
    
    for u in usuarios_inactivos:
        if u in sesiones_activas:
            del sesiones_activas[u]

    return render_template('control_sesiones.html', sesiones=sesiones_activas)

@app.route('/cerrar_sesion_remota/<usuario>', methods=['POST'])
def cerrar_sesion_remota(usuario):
    if not es_admin():
        return redirect('/')
    
    if usuario in sesiones_activas:
        del sesiones_activas[usuario]
        flash(f"La sesión del usuario {usuario} ha sido cerrada remotamente.", "success")
    else:
        flash(f"El usuario {usuario} no tiene una sesión activa.", "warning")
        
    return redirect('/ver_sesiones')

if __name__ == '__main__':
    init_db()
    from waitress import serve
    print("=====================================================")
    print(" SERVIDOR EL CARDENAL INICIADO ")
    print(" Puerto: 5004")
    print("=====================================================")
    serve(app, host='0.0.0.0', port=5004)